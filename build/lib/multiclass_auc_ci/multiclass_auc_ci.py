import openpyxl
import numpy as np
from .calc_ci import rounding
from .all_type_ci import (
    macro_normal,
    weighted_normal,
    micro_normal,
    handtill_normal,
    macro_percentile,
    weighted_percentile,
    micro_percentile,
    handtill_percentile
)


def auc_ci_select(auc_type,confidence_type,resample_num,label,score,random_seed,alpha):
    if (auc_type == "macro" and confidence_type == "normal"):
        lower,upper = macro_normal(label,score,resample_num,random_seed,alpha)

    elif (auc_type == "weighted" and confidence_type == "normal"):
        lower,upper = weighted_normal(label,score,resample_num,random_seed,alpha)

    elif (auc_type == "micro" and confidence_type == "normal"):
        lower,upper = micro_normal(label,score,resample_num,random_seed,alpha)

    elif (auc_type == "handtill" and confidence_type == "normal"):
        lower,upper = handtill_normal(label,score,resample_num,random_seed,alpha)

    elif (auc_type == "macro" and confidence_type == "percentile"):
        lower,upper = macro_percentile(label,score,resample_num,random_seed,alpha)
        
    elif (auc_type == "weighted" and confidence_type == "percentile"):
        lower,upper = weighted_percentile(label,score,resample_num,random_seed,alpha)

    elif (auc_type == "micro" and confidence_type == "percentile"):
        lower,upper = micro_percentile(label,score,resample_num,random_seed,alpha)

    elif (auc_type == "handtill" and confidence_type == "percentile"):
        lower,upper = handtill_percentile(label,score,resample_num,random_seed,alpha)

    else:
        print("Error:Please enter the correct name for auc_type or confidence_type")
        return


    return lower,upper

def multiclass_auc_ci(
    auc_type, #"macro" or "weighted" or "micro" or "handtill"
    confidence_type, #"normal" or "percentile"
    alpha, #0～1
    resample_num,
    example = None, #1～12
    label = None,
    score = None,
    random_seed = None,
    digit = 0.0001
):
    if example is None and (score is None or label is None):
        print("Error:No score or label has been entered")
        return
    
    elif example is not None:
        n_file = 108
        n_class = 3
        example_label = np.zeros(n_file)
        example_score = np.zeros((n_file,n_class))
    
        if example < 1 or 12 < example :
            print("Error:Please enter a value from 1 to 12 for example")
            return
        elif 0 < example and example < 7:
            data_file = openpyxl.load_workbook(f'multiclass_auc_ci/data/withAI/withAI_r{example}.xlsx',data_only = True)
            data_sheet = data_file["Sheet1"]
        elif 6 < example and example < 12:
            data_file = openpyxl.load_workbook(f'multiclass_auc_ci/data/withoutAI/withoutAI_r{example - 6}.xlsx',data_only = True)
            data_sheet = data_file["Sheet1"]

        for case_no in range(2,n_file+2):
            example_label[case_no-2] = data_sheet.cell(row=case_no,column=4).value
            for class_no in range(n_class):
                example_score[case_no-2][class_no] = data_sheet.cell(row=case_no,column=20+class_no).value

        lower,upper = auc_ci_select(auc_type,confidence_type,resample_num,example_label,example_score,random_seed,alpha)

    else:
        lower,upper = auc_ci_select(auc_type,confidence_type,resample_num,label,score,random_seed,alpha)

    print(rounding(lower,digit),rounding(upper,digit))
    return rounding(lower,digit),rounding(upper,digit)

multiclass_auc_ci("macro","normal",0.95,10,example=7)